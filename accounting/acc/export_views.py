from django.shortcuts import render, redirect
from openpyxl import Workbook
from openpyxl.styles import Font, Border, NamedStyle, Side, Alignment, PatternFill, numbers, NumberFormatDescriptor
from openpyxl.utils import get_column_letter
from .models import *
from django.db.models import Q, Sum
from datetime import date, datetime
from .models import Export_Data_xls
from django.contrib.auth.decorators import login_required


@login_required
def make_export(request):
    ctx = {}

    ctx['exports'] = Export_Data_xls.objects.all()
    if request.method == 'POST':
        json_data = make_json()

        if request.POST['xls_cell_choice'] == 'formula':
            write_excel(json_data, request, formula=True)
        elif request.POST['xls_cell_choice'] == 'value':
            write_excel(json_data, request, formula=False)

    return render(request, 'export/export_history.html', ctx)


def make_json():
    base_customers_var = Customers.objects.all()
    json_data = []

    for customer in base_customers_var:
        water_values = None
        sum_pay_water = None
        last_pay_water = None
        water_debth = None
        electro_values_day = None
        electro_values_night = None
        sum_pay_el_day = None
        sum_pay_el_night = None
        last_pay_el = None
        electro_day_debth = None
        electro_night_debth = None
        surname = customer.surname if customer.surname else ''
        name = customer.name if customer.name else ''
        second_name = customer.second_name if customer.second_name else ''
        data_dict = {}
        data_dict.update(
            {
                'fio': f'{surname} {name} {second_name}',
                'adress': f'{customer.massive}-{customer.line}-{customer.sector}',
                'phone_1': customer.phone_1,
                # 'phone_2': customer.phone_2,
                # 'email': customer.email,
            })
        try:
            water_values = Water_values.objects.filter(customer=customer).latest('add_date')
            data_dict.update(
                {'last_water_value': water_values.w_values,
                 'last_add_date_water_value': datetime.date(water_values.add_date),
                 })

        except Exception as e:
            data_dict.update({
                'last_water_value': None,
                'last_add_date_water_value': None})
        try:
            sum_pay_water = Payment_Data.objects.filter(customer=customer).aggregate(Sum('payment_w_values'))
            sum_pay_water = sum_pay_water['payment_w_values__sum']
            last_pay_water = (
                Payment_Data.objects.filter(customer=customer, payment_w_values__gt=0).latest('payment_date_check'))
            data_dict.update(
                {'sum_pay_water': sum_pay_water,
                 'last_pay_water': last_pay_water.payment_date_check})

        except Exception as e:
            data_dict.update({
                'sum_pay_water': None,
                'last_pay_water': None})
        try:
            water_debth = water_values.w_values - sum_pay_water
            data_dict.update({
                'water_debth': water_debth})
        except Exception as e:
            data_dict.update({
                'water_debth': None})

        try:
            electro_values_day = Electro_values.objects.filter(customer=customer, e_values_daytime__isnull=False) \
                .latest('add_date')
            electro_values_night = Electro_values.objects.filter(customer=customer, e_values_nighttime__isnull=False) \
                .latest('add_date')

            data_dict.update({
                'last_electro_values_day': electro_values_day.e_values_daytime,
                'last_electro_values_night': electro_values_night.e_values_nighttime,
                'last_add_electro_values': datetime.date(electro_values_day.add_date),
            })
        except Exception as e:
            data_dict.update({
                'last_electro_values_day': None,
                'last_electro_values_night': None,
                'last_add_electro_values': None})

        try:
            sum_pay_el_day = Payment_Data.objects.filter(customer=customer).aggregate(Sum('payment_e_values_daytime'))
            sum_pay_el_day = sum_pay_el_day['payment_e_values_daytime__sum']
            last_pay_el = Payment_Data.objects.filter(Q(payment_e_values_daytime__gt=0) |
                                                      Q(payment_e_values_nighttime__gt=0), customer=customer).latest(
                'payment_date_check')
            sum_pay_el_night = Payment_Data.objects.filter(customer=customer).aggregate(
                Sum('payment_e_values_nighttime'))
            sum_pay_el_night = sum_pay_el_night['payment_e_values_nighttime__sum']

            data_dict.update({
                'sum_pay_el_day': sum_pay_el_day,
                'sum_pay_el_night': sum_pay_el_night,
                'last_pay_el': last_pay_el.payment_date_check,
            })
        except Exception as e:

            data_dict.update({
                'sum_pay_el_day': None,
                'sum_pay_el_night': None,
                'last_pay_el': None,
            })
        try:
            electro_day_debth = electro_values_day.e_values_daytime - sum_pay_el_day
            electro_night_debth = electro_values_night.e_values_nighttime - sum_pay_el_night
            data_dict.update({
                'electro_day_debth': electro_day_debth,
                'electro_night_debth': electro_night_debth, })
        except Exception as e:
            data_dict.update({
                'electro_day_debth': None,
                'electro_night_debth': None})

        json_data.append(data_dict)
    return json_data


def write_excel(json_data, request, formula):
    keys = json_data[0].keys()  # получаем все ключи словаря
    keys = [key for key in keys]  # помещаем все ключи словаря в список для последующих итераций

    wb = Workbook()
    date_today = datetime.now().strftime("%Y_%m_%d_%H-%M")
    date_today_format = datetime.now().strftime("%Y.%m.%d")
    wb.create_sheet(title=date_today, index=0)
    wb.remove(wb['Sheet'])
    sheet = wb[date_today]

    # Пробуем создать именованый стиль
    # Стиль для тела таблицы
    body_style_1 = NamedStyle(name='body_1')
    body_style_1.font = Font(bold=False, size=11, name='Calibri')
    border = Side(style='thin', color='000000')
    body_style_1.border = Border(left=border, right=border, top=border, bottom=border)
    body_style_1.alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)
    body_style_1.fill = PatternFill(fill_type='solid', start_color='EDEDED')
    wb.add_named_style(body_style_1)

    body_alert = NamedStyle(name='body_alert')
    body_alert.fill = PatternFill(fill_type='solid', start_color='ffcccc')
    wb.add_named_style(body_alert)

    # Стиль для шапки таблицы
    head_style_1 = NamedStyle(name='head_1')
    head_style_1.font = Font(bold=True, size=14, name='Times New Roman')
    border = Side(style='thin', color='000000')
    head_style_1.border = Border(left=border, right=border, top=border, bottom=border)
    head_style_1.fill = PatternFill(fill_type='solid', start_color='ffff00')
    head_style_1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wb.add_named_style(head_style_1)

    head_style_2 = NamedStyle(name='head_2')
    head_style_2.font = Font(bold=True, size=11, name='Times New Roman', color='ffffff')
    border = Side(style='thin', color='ffffff')
    head_style_2.border = Border(left=border, right=border, top=border, bottom=border)
    head_style_2.fill = PatternFill(fill_type='solid', start_color='4f81bd')
    head_style_2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wb.add_named_style(head_style_2)

    sheet.merge_cells('A1:D1')
    sheet['A1'].value = f'Отчет на дату: {date_today_format}'
    sheet['A1'].style = 'head_1'
    sheet['A1'].alignment = Alignment(horizontal='general', vertical='center', wrap_text=True)

    sheet.merge_cells('E1:I1')
    sheet['E1'].value = 'ВОДА'
    sheet['E1'].style = 'head_1'

    sheet.merge_cells('J1:Q1')
    sheet['J1'].value = 'ЭЛЕКТРИЧЕСТВО'
    sheet['J1'].style = 'head_1'

    header_data = ['№', 'ФИО', 'Адрес', 'Телефон', 'Послед. показ., м3', 'Дата посл. показ.',
                   'Сумм. оплачено, м3', 'Дата посл. оплаты', 'Долг, м3',
                   'Послед. показ.(День)', 'Послед. показ.(Ночь)', 'Дата посл. показ.',
                   'Сумм. опл. (День)', 'Сумм. опл. (Ночь)', 'Дата посл. опл.', 'Долг (День), КВт',
                   'Долг (Ночь), КВт',
                   ]
    col_count = 1
    # устанавливаем ширину столбцов
    for field in header_data:
        cell = sheet.cell(row=2, column=col_count)
        cell.value = field
        cell.style = 'head_2'
        width_12 = [6, 8, 9, 12, 15, 16, 17]
        if col_count == 1:
            sheet.column_dimensions[get_column_letter(col_count)].width = 5
        elif col_count == 2:
            sheet.column_dimensions[get_column_letter(col_count)].width = 30
        elif col_count == 3 or col_count == 5:
            sheet.column_dimensions[get_column_letter(col_count)].width = 10
        elif col_count == 7 or col_count == 13 or col_count == 14:
            sheet.column_dimensions[get_column_letter(col_count)].width = 14
        elif col_count in width_12:
            sheet.column_dimensions[get_column_letter(col_count)].width = 12

        else:
            sheet.column_dimensions[get_column_letter(col_count)].width = 15

        col_count += 1
    sheet.auto_filter.ref = 'A2:Q2'

    def debth_formula(j, dict_elem_count):
        # Для колонки "Долг, вода" устанавливаем формулу для подсчета и подкрашиваем ячейки.

        if j == 7:
            cell.value = f'={get_column_letter(j - 2)}{dict_elem_count}-{get_column_letter(j)}{dict_elem_count}'
        elif j == 14 or j == 15:
            cell.value = f'={get_column_letter(j - 4)}{dict_elem_count}-{get_column_letter(j - 1)}{dict_elem_count}'

    def debth_value(j,elem, keys):
        try:
            if j == 7:
                if elem[keys[j - 4]] is None:
                    last_value = 0
                else:
                    last_value = elem[keys[j - 4]]
                if elem[keys[j - 2]] is None:
                    pay = 0
                else:
                    pay = elem[keys[j - 2]]
            elif j == 14 or j == 15:
                if elem[keys[j - 6]] is None:
                    last_value = 0
                else:
                    last_value = elem[keys[j - 6]]
                if elem[keys[j - 3]] is None:
                    pay = 0
                else:
                    pay = elem[keys[j - 3]]
            debth = last_value - pay
            cell.value = debth
        except Exception as e:
            print(e)

    def debth_color(j, elem, keys):
        # Подкрашивание для должников - красным и зеленым для заплативших вперед.
        cell.style = 'body_1'
        try:
            if j == 7:
                if elem[keys[j - 4]] is None:
                    last_value = 0
                else:
                    last_value = elem[keys[j - 4]]
                if elem[keys[j - 2]] is None:
                    pay = 0
                else:
                    pay = elem[keys[j - 2]]
            elif j == 14 or j == 15:
                if elem[keys[j - 6]] is None:
                    last_value = 0
                else:
                    last_value = elem[keys[j - 6]]
                if elem[keys[j - 3]] is None:
                    pay = 0
                else:
                    pay = elem[keys[j - 3]]
            debth = last_value - pay
            if debth > 0:
                cell.fill = PatternFill(fill_type='solid', start_color='ffcccc')
            elif debth < 0:
                cell.fill = PatternFill(fill_type='solid', start_color='66ff99')
        except Exception as e:
            print(e)

    dict_elem_count = 3
    for elem in json_data:
        cell = sheet.cell(row=dict_elem_count, column=1)
        cell.value = dict_elem_count - 2
        cell.style = 'body_1'
        for j in range(len(keys)):
            cell = sheet.cell(row=dict_elem_count, column=j + 2)

            # Для колонок с долгом делаем: 1. устанавливаем формулу; 2.Подкрашиваем
            if j == 7 or j == 14 or j == 15:
                if formula == True:
                    debth_formula(j, dict_elem_count)
                else:
                    debth_value(j,elem, keys)
                debth_color(j,elem, keys)
            else:
                cell.value = elem[keys[j]]
                cell.style = 'body_1'
            # Определяем формат Дата для ячеек
            if j == 4 or j == 6 or j == 10 or j == 13:
                cell.number_format = numbers.FORMAT_DATE_XLSX14

        dict_elem_count += 1

    sheet.row_dimensions[1].height = 20
    sheet.row_dimensions[2].height = 30


    filename = f'Показания-{date_today}.xlsx'

    wb.save(filename=f'media/export_xls/{filename}')

    # Добавление в базу данных о экспорте:
    add_export = Export_Data_xls(user=request.user.username, export_file=filename)
    add_export.save()
